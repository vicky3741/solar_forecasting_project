"""
=========================================================
Solar Forecasting Project
Single-Sheet Schedule Report
=========================================================
One sheet holding the entire workflow, as the evaluation
document describes it: every 15-minute block down the rows,
every scheduling time across the columns.

Each column shows what THAT scheduling run published for
that block. A run's column is blank for blocks that had
already passed when it ran - the document's rule ("the
schedule values for the blocks that have already passed
will remain unchanged") made visible as a staircase. The
final published value, the actual meter reading and the
error close the table.

Written for speed: plain computed values rather than
formulas, one merged range, and a handful of shared style
objects reused across every cell. A workbook whose formulas
carry no cached result forces a full recalculation the
moment it opens, which is what made an earlier version feel
sluggish in phone and browser viewers.

Run:  python -m tests.build_single_sheet_report [YYYY-MM-DD]
=========================================================
"""

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DAY = sys.argv[1] if len(sys.argv) > 1 else "2026-07-25"
SRC = Path("outputs/schedules")
OUT = Path(f"outputs/reports/Schedule_{DAY}_single_sheet.xlsx")

CAPACITY_MW = 5.1
PLANT = "Sirmour Solar Plant"
BLOCK_HOURS = 0.25          # each scheduling block is 15 minutes

FONT = "Arial"
NAVY = "1F3864"
BLUE = "DCE6F1"
GREY = "F2F2F2"
AMBER = "FCE4D6"

# --- shared style objects: created once, referenced everywhere ---
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

_thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

F_TITLE = Font(name=FONT, size=13, bold=True, color=NAVY)
F_SUB = Font(name=FONT, size=9, italic=True, color="595959")
F_HEAD = Font(name=FONT, size=9.5, bold=True, color="FFFFFF")
F_BODY = Font(name=FONT, size=9.5)
F_OWNER = Font(name=FONT, size=9.5, bold=True)
F_POS = Font(name=FONT, size=9.5, color="C00000")
F_NEG = Font(name=FONT, size=9.5, color="1F4E79")
F_SECTION = Font(name=FONT, size=11, bold=True, color=NAVY)
F_LABEL = Font(name=FONT, size=10, bold=True)
F_VALUE = Font(name=FONT, size=10)
F_NOTE = Font(name=FONT, size=9, color="595959")

FILL_HEAD = PatternFill("solid", fgColor=NAVY)
FILL_OWNER = PatternFill("solid", fgColor=BLUE)
FILL_GREY = PatternFill("solid", fgColor=GREY)
FILL_AMBER = PatternFill("solid", fgColor=AMBER)


def main():

    schedule = pd.read_csv(SRC / f"day_schedule_{DAY}.csv")
    per_run = pd.read_csv(SRC / f"day_schedule_{DAY}_per_run.csv")
    run_log = pd.read_csv(SRC / f"day_schedule_{DAY}_run_log.csv")

    run_times = list(run_log["scheduling_time"])

    matrix = per_run.pivot_table(
        index="block", columns="scheduling_time",
        values="scheduled_mw", aggfunc="first",
    )

    base = schedule.set_index("block")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Schedule {DAY}"

    n_run = len(run_times)
    total_cols = 2 + n_run + 4

    # ---------------- title (the only merged range) ----------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1,
                value=f"{PLANT} - Reconstructed Day Schedule - {DAY}   "
                      f"(all values in MW, installed capacity {CAPACITY_MW} MW)")
    c.font = F_TITLE
    c.alignment = CENTER
    ws.row_dimensions[1].height = 20

    c = ws.cell(row=2, column=1,
                value="Each scheduling column shows what that run scheduled for the block. "
                      "Blank = the block had already passed when that run executed, so it "
                      "kept the value an earlier run gave it.")
    c.font = F_SUB

    # ---------------- header ----------------
    HEAD = 4
    labels = (["Block", "Time"] + run_times
              + ["Final (MW)", "Actual (MW)", "Error (MW)", "Set by run"])

    for i, label in enumerate(labels, start=1):
        cell = ws.cell(row=HEAD, column=i, value=label)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        cell.alignment = CENTER_WRAP
        cell.border = BOX
    ws.row_dimensions[HEAD].height = 22

    first_run_col = 3
    col_final = 2 + n_run + 1
    col_actual = col_final + 1
    col_error = col_actual + 1
    col_src = col_error + 1

    # ---------------- data ----------------
    start = HEAD + 1
    blocks = list(base.index)

    for r, block in enumerate(blocks, start=start):

        rec = base.loc[block]
        unscored = not bool(rec["actual_is_real"])

        ws.cell(row=r, column=1, value=int(block))
        ws.cell(row=r, column=2, value=rec["block_time"])

        for i, rt in enumerate(run_times):
            value = matrix.loc[block, rt] if rt in matrix.columns else None
            if pd.isna(value):
                value = None
            cell = ws.cell(row=r, column=first_run_col + i, value=value)
            if value is not None:
                cell.number_format = "0.0000"
                if rt == rec["scheduled_at"]:
                    cell.fill = FILL_OWNER
                    cell.font = F_OWNER

        ws.cell(row=r, column=col_final,
                value=float(rec["scheduled_mw"])).number_format = "0.0000"

        actual = None if pd.isna(rec["actual_mw"]) else float(rec["actual_mw"])
        error = None if pd.isna(rec["error_mw"]) else float(rec["error_mw"])

        ws.cell(row=r, column=col_actual, value=actual).number_format = "0.0000"

        ec = ws.cell(row=r, column=col_error, value=error)
        ec.number_format = "0.0000;-0.0000"
        if error is not None:
            ec.font = F_POS if error > 0 else F_NEG

        ws.cell(row=r, column=col_src, value=rec["scheduled_at"])

        for col in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = BOX
            cell.alignment = CENTER
            if cell.font is None or cell.font.sz is None:
                cell.font = F_BODY
            if unscored:
                cell.fill = FILL_AMBER

    last_row = start + len(blocks) - 1

    # ---------------- totals, computed here (no formulas) ----------------
    # Per-block values are power (MW) and every one of them must sit below
    # the 5.1 MW rating. Power CANNOT be summed - adding 45 block values
    # produced 40.2 "MW", which is meaningless and above plant capacity
    # (Vikrant caught this on 2026-07-27). A total has to be ENERGY:
    # MW x 0.25 h per block = MWh. So the sheet reports the average and the
    # peak in MW, where the capacity check applies, and the total in MWh,
    # where it does not.
    scored = schedule[schedule["actual_is_real"] & schedule["actual_mw"].notna()]

    avg_predicted = float(scored["scheduled_mw"].mean())
    avg_actual = float(scored["actual_mw"].mean())
    peak_predicted = float(scored["scheduled_mw"].max())
    peak_actual = float(scored["actual_mw"].max())

    # Arithmetic column totals, requested so the MW columns can be
    # checked by adding them up. Not a power level - no instant of the
    # day reached this - and exactly 4x the MWh figure, since each block
    # is a quarter hour.
    sum_predicted = float(scored["scheduled_mw"].sum())
    sum_actual = float(scored["actual_mw"].sum())
    sum_error = sum_predicted - sum_actual
    sum_abs_error = float(scored["error_mw"].abs().sum())

    energy_predicted = float(scored["scheduled_mw"].sum() * BLOCK_HOURS)
    energy_actual = float(scored["actual_mw"].sum() * BLOCK_HOURS)
    energy_error = energy_predicted - energy_actual
    energy_abs_error = float(scored["error_mw"].abs().sum() * BLOCK_HOURS)

    mae = float(scored["error_mw"].abs().mean())
    rmse = float(np.sqrt((scored["error_mw"] ** 2).mean()))
    deviation = mae / CAPACITY_MW * 100

    # ---------------- summary rows closing the block table ----------------
    # Two rows, because MW and MWh answer different questions: the AVERAGE
    # stays in MW and can be checked against the 5.1 MW rating; the TOTAL
    # must be MWh, since power cannot be added up.
    summary_rows = [
        ("AVERAGE", "MW", avg_predicted, avg_actual, avg_predicted - avg_actual, "0.0000"),
        ("SUM", "MW", sum_predicted, sum_actual, sum_error, "0.0000"),
        ("TOTAL", "MWh", energy_predicted, energy_actual, energy_error, "0.000"),
    ]

    trow = last_row
    for label, unit, predicted, actual, error, fmt in summary_rows:

        trow += 1

        for col, value, number_format in (
            (1, label, None),
            (2, f"{len(scored)} blk", None),
            (col_final, round(predicted, 4), fmt),
            (col_actual, round(actual, 4), fmt),
            (col_error, round(error, 4), fmt if error >= 0 else fmt),
            (col_src, unit, None),
        ):
            cell = ws.cell(row=trow, column=col, value=value)
            cell.font = F_HEAD
            cell.fill = FILL_HEAD
            cell.alignment = CENTER
            cell.border = BOX
            if number_format:
                cell.number_format = number_format

        # Run columns stay blank: each run covers a different number of
        # blocks, so its column figure would compare against nothing.
        for i in range(n_run):
            cell = ws.cell(row=trow, column=first_run_col + i)
            cell.fill = FILL_HEAD
            cell.border = BOX

    ws.cell(row=trow + 1, column=1,
            value=f"All three rows cover the same {len(scored)} blocks with a real meter "
                  f"reading. AVERAGE (MW) is the day's mean output and sits below the "
                  f"{CAPACITY_MW} MW rating. SUM (MW) is the arithmetic column total - "
                  "useful for checking the columns add up, but it is not a power level, "
                  "since no instant of the day reached it. TOTAL (MWh) is the electricity "
                  "actually generated; each block is 15 minutes, so MWh = SUM x 0.25."
            ).font = F_NOTE

    mrow = trow + 3
    ws.cell(row=mrow, column=1, value="ACCURACY vs ACTUAL METER DATA").font = F_SECTION

    rows = [
        ("Blocks scored (real meter readings)", len(scored), "0"),
        (f"--- POWER, in MW (plant rating {CAPACITY_MW} MW) ---", None, None),
        ("Average predicted (MW)", round(avg_predicted, 4), "0.0000"),
        ("Average actual (MW)", round(avg_actual, 4), "0.0000"),
        ("Peak predicted (MW)", round(peak_predicted, 4), "0.0000"),
        ("Peak actual (MW)", round(peak_actual, 4), "0.0000"),
        ("Mean absolute error (MW)", round(mae, 4), "0.0000"),
        ("Root mean squared error (MW)", round(rmse, 4), "0.0000"),
        ("Average percentage deviation (%)", round(deviation, 2), "0.00"),
        ("Sum of predicted blocks (MW)", round(sum_predicted, 4), "0.0000"),
        ("Sum of actual blocks (MW)", round(sum_actual, 4), "0.0000"),
        ("Sum of error - predicted minus actual (MW)", round(sum_error, 4),
         "+0.0000;-0.0000"),
        ("Sum of absolute error (MW)", round(sum_abs_error, 4), "0.0000"),
        ("--- ENERGY, in MWh (SUM above x 0.25 h per block) ---", None, None),
        ("Total predicted energy (MWh)", round(energy_predicted, 3), "0.000"),
        ("Total actual energy (MWh)", round(energy_actual, 3), "0.000"),
        ("Total error - predicted minus actual (MWh)", round(energy_error, 3),
         "+0.000;-0.000"),
        ("Total absolute error (MWh)", round(energy_abs_error, 3), "0.000"),
    ]

    for j, (label, value, fmt) in enumerate(rows, start=1):

        lc = ws.cell(row=mrow + j, column=1, value=label)

        if value is None:                 # a section divider, not a metric
            lc.font = F_SUB
            continue

        lc.font = F_LABEL
        vc = ws.cell(row=mrow + j, column=4, value=value)
        vc.font = F_VALUE
        vc.number_format = fmt
        vc.fill = FILL_GREY
        vc.border = BOX
        vc.alignment = CENTER

    # ---------------- legend ----------------
    lrow = mrow + len(rows) + 2
    notes = [
        ("Method", "Per 'Schedule Generation Workflow for Model Evaluation': each scheduling "
                   "time uses only the Windy clip stored at that time and meter data up to "
                   "block T, then schedules to end of day. Past blocks stay frozen; only "
                   "future blocks are rewritten."),
        ("Blue cells", "The scheduling run whose value was finally published for that block."),
        ("Blank cells", "That block had already passed when the run executed."),
        ("Orange rows", "No measured meter reading - excluded from every figure above."),
        ("Deviation", f"Mean absolute error as a percentage of the {CAPACITY_MW} MW "
                      "installed capacity."),
        ("MW vs MWh", "MW is a rate, like speed - every block value must sit below the "
                      f"{CAPACITY_MW} MW rating. MWh is a quantity, like distance, and is "
                      "the only correct way to express a day total. Driving at 60 km/h for "
                      "4 hours gives 240 km, not 240 km/h."),
        ("Total error", "Predicted minus actual. Morning under-forecasts and afternoon "
                        "over-forecasts partly cancel here, so the absolute total is the "
                        "stricter measure."),
    ]
    for i, (k, v) in enumerate(notes):
        ws.cell(row=lrow + i, column=1, value=k).font = F_LABEL
        cell = ws.cell(row=lrow + i, column=3, value=v)
        cell.font = F_NOTE
        cell.alignment = LEFT_WRAP
        ws.row_dimensions[lrow + i].height = 24

    # ---------------- widths ----------------
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 9
    for i in range(n_run):
        ws.column_dimensions[get_column_letter(first_run_col + i)].width = 10.5
    for col in (col_final, col_actual, col_error, col_src):
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws.freeze_panes = ws.cell(row=start, column=3)

    OUT.parent.mkdir(parents=True, exist_ok=True)

    # Excel holds an exclusive lock on an open workbook, so a rebuild
    # while the previous version is still open would otherwise die with
    # PermissionError. Fall back to a numbered name rather than losing
    # the run.
    target = OUT
    for attempt in range(1, 20):
        try:
            wb.save(target)
            break
        except PermissionError:
            target = OUT.with_name(f"{OUT.stem}_{attempt}{OUT.suffix}")
    else:
        raise PermissionError(f"Could not write {OUT} or any fallback name")

    if target != OUT:
        print(f"NOTE: {OUT.name} is open in Excel - saved as {target.name} instead")

    print(f"written: {target}")
    print(f"one sheet '{ws.title}': {len(blocks)} blocks x {n_run} scheduling times")
    print(f"  blocks scored          : {len(scored)}")
    print(f"  POWER  avg predicted   : {avg_predicted:.4f} MW   "
          f"(peak {peak_predicted:.4f} MW)")
    print(f"  POWER  avg actual      : {avg_actual:.4f} MW   "
          f"(peak {peak_actual:.4f} MW)")
    print(f"  POWER  MAE / RMSE      : {mae:.4f} / {rmse:.4f} MW")
    print(f"  POWER  deviation       : {deviation:.2f}%")
    print(f"  SUM    predicted       : {sum_predicted:.4f} MW  (column total)")
    print(f"  SUM    actual          : {sum_actual:.4f} MW  (column total)")
    print(f"  SUM    error           : {sum_error:+.4f} MW")
    print(f"  ENERGY total predicted : {energy_predicted:.3f} MWh")
    print(f"  ENERGY total actual    : {energy_actual:.3f} MWh")
    print(f"  ENERGY total error     : {energy_error:+.3f} MWh")
    print(f"  ENERGY absolute error  : {energy_abs_error:.3f} MWh")
    print()
    print(f"  sanity: every MW figure below plant rating {CAPACITY_MW} MW -> "
          f"{max(peak_predicted, peak_actual) <= CAPACITY_MW}")


if __name__ == "__main__":
    main()
