"""
=========================================================
Solar Forecasting Project
Frozen Enercast xlsx -> data/enercast CSV
=========================================================
MENTOR GUIDANCE 2026-08-06: use the "frozen" Enercast export
as the comparison baseline, not a single day-ahead file.

Enercast re-issues its schedule through the day. The frozen
export resolves, for every block, the value from the LAST
revision that became effective BEFORE that block went live
(Arrival Time + ~1.5 h = Effective Time in the source), so it
is what Enercast actually had committed at gate closure -
never a later revision that already saw the weather turn.

This matters because our own pipeline re-forecasts 7x/day.
Grading our intraday forecast against a plain day-ahead
Enercast file flatters us; the frozen file is the like-for-
like baseline. On 2026-08-06 the switch cut Enercast's DSM
penalty from Rs 2859 to Rs 1515 - our own numbers do not move
at all, since we are graded against actual meter data.

Only Block / Time / Scheduled MW are written, matching what
Evaluator.load_enercast_day expects; the revision/arrival
provenance columns stay in the source xlsx.

Run:  python -m tests.convert_frozen_enercast <frozen.xlsx> <YYYY-MM-DD>
=========================================================
"""

import argparse
from pathlib import Path

import openpyxl
import pandas as pd

from config.config import settings


def parse_args():

    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", help="the frozen Enercast export")
    parser.add_argument("day", help="YYYY-MM-DD the export covers")
    return parser.parse_args()


def read_frozen(path):
    """Block / Time / Scheduled MW out of the frozen export."""

    worksheet = openpyxl.load_workbook(path, data_only=True).active

    header = [cell.value for cell in worksheet[1]]
    for required in ("Block", "Time", "Scheduled MW"):
        if required not in header:
            raise ValueError(f"{path}: missing required column {required!r}")

    block_col = header.index("Block")
    time_col = header.index("Time")
    mw_col = header.index("Scheduled MW")

    records = []
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):

        block = row[block_col].value
        if block is None:
            continue

        # The export writes every cell as text, so Block and Scheduled MW
        # both arrive as strings and must be cast, not trusted.
        records.append({
            "Block": int(block),
            "Time": str(row[time_col].value),
            "Scheduled MW": float(row[mw_col].value),
        })

    return pd.DataFrame(records).sort_values("Block").reset_index(drop=True)


def main():

    args = parse_args()
    frame = read_frozen(args.xlsx)

    day = pd.Timestamp(args.day)
    filename = f"Sirmour_{day.day}{day.strftime('%B').lower()}_enercast.csv"
    out_path = Path(settings["paths"]["enercast_data"]) / filename

    if len(frame) != 96:
        print(f"WARNING: {len(frame)} blocks, expected 96 for a full day")

    previous = None
    if out_path.exists():
        previous = pd.read_csv(out_path)

    frame.to_csv(out_path, index=False)

    print(f"blocks written: {len(frame)}")
    print(f"saved: {out_path}")

    if previous is not None:
        merged = previous.merge(frame, on="Block", suffixes=("_old", "_new"))
        changed = merged[
            (merged["Scheduled MW_old"] - merged["Scheduled MW_new"]).abs() > 1e-9
        ]
        print(f"replaced an existing file - {len(changed)} of {len(merged)} blocks changed")


if __name__ == "__main__":
    main()
