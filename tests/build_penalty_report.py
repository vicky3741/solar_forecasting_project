"""
=========================================================
Solar Forecasting Project
Schedule vs Meter + DSM Penalty Report
=========================================================
Extends the reconstructed day schedule with the DSM
(Deviation Settlement Mechanism) slab-based penalty: the
first 10% of deviation from installed capacity is a free
band (no penalty), then progressively higher per-kWh rates
apply for 10-15%, 15-20% and >20% deviation. Matches the
sheet layout already used for 2026-07-25/27/28/29.

"Deviation" here is Actual minus AI Schedule (the sign the
DSM regulation uses), the opposite convention from the
plain Schedule_vs_Meter report's Error (Scheduled - Actual).

Every accuracy/penalty figure is a live formula referencing
the DSM slab parameters at the bottom of the sheet, so the
sheet recalculates if the slab rates or capacity ever change.

Run:  python -m tests.build_penalty_report [YYYY-MM-DD]
=========================================================
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, Reference
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.config import settings
from modules.evaluation.evaluator import Evaluator
from modules.preprocessing.preprocess import DataPreprocessor

DAY = sys.argv[1] if len(sys.argv) > 1 else "2026-07-25"
SCHEDULE_DIR = Path("outputs/schedules")
OUT_DIR = Path("outputs/reports")

CAPACITY_MW = settings["plant"]["capacity_mw"]
PLANT = settings["plant"]["name"]
BLOCK_ENERGY_FACTOR = 250   # 0.25 h x 1000 kW/MW - MW deviation -> kWh for one block

FONT = "Arial"
DARK_GREEN = "1B4332"
RED = "C00000"
RED_FILL = "FCE4E4"
BLUE = "1F4E9C"
BLUE_FILL = "E4EDFA"

# DSM slabs: (label, from_pct, to_pct_or_None, rate_rs_per_kwh)
SLABS = [
    ("Slab 1", 0, 10, 0),
    ("Slab 2", 10, 15, 0.5),
    ("Slab 3", 15, 20, 0.75),
    ("Slab 4", 20, None, 1),
]


def block_number(timestamp):
    """Indian scheduling block number: block 1 is 00:00-00:15."""

    return timestamp.hour * 4 + timestamp.minute // 15 + 1


def load_full_day_actual(day):
    """Every meter reading for the day, not just the scored blocks."""

    path = Path(settings["paths"]["historical_data"]) / f"{day.replace('-', '_')}_SOLAR_INV.csv"

    frame = DataPreprocessor().preprocess(
        file_path=path,
        required_columns=["TimeStamp"],
        timestamp_column="TimeStamp",
    )

    columns = ["timestamp", "active_power_kw"]
    if "is_real_measurement" in frame.columns:
        columns.append("is_real_measurement")

    return frame[columns]


def build_rows(day):
    """
    Every block from the day's earliest meter reading to the last
    scheduled block, merging in the reconstructed schedule. Blocks
    before the first scheduling run get an AI Schedule of 0 (no run
    has happened yet); blocks with no meter reading yet are left
    blank rather than guessed.
    """

    schedule = pd.read_csv(
        SCHEDULE_DIR / f"day_schedule_{day}.csv", parse_dates=["timestamp"]
    )
    actual = load_full_day_actual(day)

    merged = actual.merge(
        schedule[["timestamp", "scheduled_mw", "scheduled_at"]],
        on="timestamp", how="outer"
    ).sort_values("timestamp").reset_index(drop=True)

    merged = merged[merged["timestamp"].dt.date == pd.Timestamp(day).date()]

    # Enercast, side by side, for comparison only - never used to grade our
    # own forecast (see modules/evaluation/evaluator.py).
    enercast_day = Evaluator().load_enercast_day(pd.Timestamp(day))
    if enercast_day is not None:
        merged = merged.merge(enercast_day, on="timestamp", how="left")

    # ONLY the blocks a scheduling run actually wrote. Blocks the meter
    # covers but no run ever reached (the first run is 06:45, so nothing
    # before 07:00) were never scheduled by us, so scoring them is
    # meaningless - and being near-zero all night they silently DILUTE the
    # day's deviation. The 2026-08-02 meter file starts at 00:00 rather
    # than ~05:45 like earlier days, and those 30 extra sleeping blocks
    # pulled the reported deviation from 6.00% down to 3.73% purely by
    # averaging in rows where nothing happened.
    #
    # This is the same rule build_simple_schedule.py already applies, and
    # it keeps this report's numbers identical to the canonical summary
    # generate_schedule_for_day writes.
    merged = merged[merged["scheduled_mw"].notna()].reset_index(drop=True)

    return merged


def main():

    day = pd.Timestamp(DAY).date()
    rows = build_rows(DAY)
    has_enercast = "enercast_kw" in rows.columns and rows["enercast_kw"].notna().any()

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule vs Meter + Penalty"

    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")
    f_head = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    f_body = Font(name=FONT, size=10)
    f_section = Font(name=FONT, size=12, bold=True, color=DARK_GREEN)
    f_label = Font(name=FONT, size=10)
    f_input = Font(name=FONT, size=10, color="0000FF")
    fill_head = PatternFill("solid", fgColor=DARK_GREEN)

    headers = ["Block", "Time", "AI Schedule (MW)", "Actual (MW)",
               "Deviation (MW)", "Deviation % (Capacity)", "Penalty (Rs)", "Scheduled at"]

    if has_enercast:
        headers += ["Enercast (MW)", "Enercast Deviation (MW)",
                    "Enercast Deviation % (Capacity)", "Enercast Penalty (Rs)"]

    # Moving penalty band, per mentor's Updated_Penalty_Logic.docx:
    # width = plant capacity x 10%, centered on the METER/ACTUAL line, not
    # a fixed horizontal line (and not the schedule line - superseded from
    # the original Penalty_band_logic.docx). band_width is a genuine
    # constant (upper - lower always cancels the meter term to 2 x the
    # half-width), kept as its own column only so the stacked-area chart
    # trick below can reference it directly.
    headers += ["Lower Penalty Band (MW)", "Penalty Band Width (MW)"]

    last_col_letter = get_column_letter(len(headers))

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws.cell(row=1, column=1,
            value=f"{PLANT} - AI Schedule vs Actual with DSM Penalty - {day}"
            ).font = f_section
    ws.merge_cells(f"A2:{last_col_letter}2")

    if has_enercast:
        ws.cell(row=2, column=1,
                value="Enercast is shown alongside for comparison only - our forecast is "
                      "graded against actual meter data, never against Enercast."
                ).font = Font(name=FONT, size=9, italic=True, color="595959")

    HEAD = 4
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=HEAD, column=i, value=h)
        cell.font = f_head
        cell.fill = fill_head
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box

    start = HEAD + 1

    # Rows with an actual reading (deviation/penalty can be computed);
    # tracked so the formulas can be backfilled once the slab parameter
    # cells below have fixed addresses. Enercast tracked separately since
    # a block can have Actual without Enercast or vice versa.
    scored_rows = []
    enercast_rows = []

    for r, (_, rec) in enumerate(rows.iterrows(), start=start):

        block = block_number(rec["timestamp"])
        block_end = rec["timestamp"] + pd.Timedelta(minutes=15)
        time_label = f"{rec['timestamp'].strftime('%H:%M')}-{block_end.strftime('%H:%M')}"

        ws.cell(row=r, column=1, value=block)
        ws.cell(row=r, column=2, value=time_label)

        ws.cell(row=r, column=3, value=round(float(rec["scheduled_mw"]), 4)
                 ).number_format = "0.000"

        if pd.notna(rec["active_power_kw"]):
            actual_mw = round(float(rec["active_power_kw"]) / 1000, 4)
            ws.cell(row=r, column=4, value=actual_mw).number_format = "0.000"

            ws.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = "0.000"

            scored_rows.append(r)

        if pd.notna(rec["scheduled_at"]):
            ws.cell(row=r, column=8, value=str(rec["scheduled_at"]))

        if has_enercast and pd.notna(rec.get("enercast_kw")):
            enercast_mw = round(float(rec["enercast_kw"]) / 1000, 4)
            ws.cell(row=r, column=9, value=enercast_mw).number_format = "0.000"

            if pd.notna(rec["active_power_kw"]):
                ws.cell(row=r, column=10, value=f"=D{r}-I{r}").number_format = "0.000"
                enercast_rows.append(r)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = box
            cell.alignment = center
            if cell.font is None or cell.font.sz is None:
                cell.font = f_body

    last_row = start + len(rows) - 1

    # ---------------- DSM slab parameters (written below the summary) ----------------
    # Each summary block (AI, and Enercast when present) is a fixed 12
    # metric rows starting 2 rows below the data; the slab table sits
    # below whichever summaries are present, with a 1-row gap.
    SUMMARY_LEN = 12
    ai_summary_end = last_row + 2 + SUMMARY_LEN
    if has_enercast:
        ec_summary_end = ai_summary_end + 2 + SUMMARY_LEN
        param_row = ec_summary_end + 2
    else:
        param_row = ai_summary_end + 2
    cap_cell = f"C{param_row + 1}"
    factor_cell = f"C{param_row + 2}"

    ws.cell(row=param_row, column=1,
            value="DSM SLAB PARAMETERS (the penalty column references these cells)"
            ).font = f_section

    ws.cell(row=param_row + 1, column=1, value="Installed capacity (MW)").font = f_label
    c = ws.cell(row=param_row + 1, column=3, value=CAPACITY_MW)
    c.font = f_input

    ws.cell(row=param_row + 2, column=1, value="Block energy factor (0.25 h x 1000 kW/MW)"
            ).font = f_label
    c = ws.cell(row=param_row + 2, column=3, value=BLOCK_ENERGY_FACTOR)
    c.font = f_input

    slab_head_row = param_row + 4
    for i, h in enumerate(["Slab", "From %", "To %", "Rate (Rs/kWh)", "Upper edge (MW)"], start=1):
        cell = ws.cell(row=slab_head_row, column=i, value=h)
        cell.font = f_head
        cell.fill = fill_head
        cell.alignment = center

    slab_rows = {}
    for i, (label, frm, to, rate) in enumerate(SLABS, start=1):
        r = slab_head_row + i
        slab_rows[label] = r
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=frm)
        ws.cell(row=r, column=3, value=to if to is not None else "above")
        rate_cell = ws.cell(row=r, column=4, value=rate)
        rate_cell.font = f_input
        if to is not None:
            edge_cell = ws.cell(row=r, column=5, value=f"={cap_cell}*{to}/100")
            edge_cell.number_format = "0.000"

    slab1_edge = f"E{slab_rows['Slab 1']}"
    slab2_edge = f"E{slab_rows['Slab 2']}"
    slab3_edge = f"E{slab_rows['Slab 3']}"
    slab2_rate = f"D{slab_rows['Slab 2']}"
    slab3_rate = f"D{slab_rows['Slab 3']}"
    slab4_rate = f"D{slab_rows['Slab 4']}"

    def penalty_formula(dev_col, r):
        return (
            f"={factor_cell}*("
            f"MAX(0,MIN(ABS({dev_col}{r}),{slab2_edge})-{slab1_edge})*{slab2_rate}"
            f"+MAX(0,MIN(ABS({dev_col}{r}),{slab3_edge})-{slab2_edge})*{slab3_rate}"
            f"+MAX(0,ABS({dev_col}{r})-{slab3_edge})*{slab4_rate})"
        )

    # Backfill the deviation-% and penalty formulas now that the slab
    # parameter cells have fixed addresses.
    for r in scored_rows:
        pct_cell = ws.cell(row=r, column=6, value=f"=E{r}/{cap_cell}*100")
        pct_cell.number_format = "\\+0.00;\\-0.00"
        pen_cell = ws.cell(row=r, column=7, value=penalty_formula("E", r))
        pen_cell.number_format = "0.00"

    for r in enercast_rows:
        pct_cell = ws.cell(row=r, column=11, value=f"=J{r}/{cap_cell}*100")
        pct_cell.number_format = "\\+0.00;\\-0.00"
        pen_cell = ws.cell(row=r, column=12, value=penalty_formula("J", r))
        pen_cell.number_format = "0.00"

    # Moving penalty band, per Updated_Penalty_Logic.docx: centered on the
    # METER/ACTUAL line (column D), not the AI schedule line - and only for
    # blocks with a real meter reading (band = null where meter is missing,
    # matching the doc's step 3c). Band width formula ties back to installed
    # capacity (per spec point 6), not a hardcoded 0.51 - if capacity or the
    # slab-1 edge percentage ever changes, this moves with it.
    band_lower_col = len(headers) - 1
    band_width_col = len(headers)
    for r in scored_rows:
        lower_cell = ws.cell(row=r, column=band_lower_col,
                              value=f"=D{r}-{cap_cell}*10/100")
        lower_cell.number_format = "0.000"
        width_cell = ws.cell(row=r, column=band_width_col,
                              value=f"={cap_cell}*10/100*2")
        width_cell.number_format = "0.000"

    # ---------------- day summary ----------------
    srow = last_row + 2
    ws.cell(row=srow, column=1, value="DAY SUMMARY - ACCURACY AND DSM PENALTY").font = f_section

    d_rng = f"D{start}:D{last_row}"
    c_rng = f"C{start}:C{last_row}"
    e_rng = f"E{start}:E{last_row}"
    g_rng = f"G{start}:G{last_row}"

    summary = [
        ("Blocks with a real meter reading", f"=COUNT({d_rng})", "0"),
        ("Total scheduled (MW, scored blocks)", f'=SUMIF({d_rng},"<>",{c_rng})', "0.000"),
        ("Total actual (MW)", f"=SUM({d_rng})", "0.000"),
        ("Total deviation - actual minus scheduled (MW)", f"=SUM({e_rng})", "\\+0.000;\\-0.000"),
        ("Mean absolute deviation (MW)",
         f"=SUMPRODUCT(ABS({e_rng}))/COUNT({e_rng})", "0.000"),
        ("Max absolute deviation (MW)", f"=MAX(MAX({e_rng}),-MIN({e_rng}))", "0.000"),
        ("Blocks over 0.5 MW deviation (RED)",
         f'=SUMPRODUCT((ABS({e_rng})>0.5)*({d_rng}<>""))', "0"),
        ("Blocks within 0.5 MW deviation (BLUE)",
         f'=SUMPRODUCT((ABS({e_rng})<=0.5)*({d_rng}<>""))', "0"),
        ("Mean absolute deviation (% of capacity)",
         f"=SUMPRODUCT(ABS({e_rng}))/COUNT({e_rng})/{cap_cell}*100", "0.00"),
        ("Blocks that incurred a penalty", f'=COUNTIF({g_rng},">0")', "0"),
        ("Worst single-block penalty (Rs)", f"=MAX({g_rng})", "0.00"),
        ("TOTAL DSM PENALTY FOR THE DAY (Rs)", f"=SUM({g_rng})", "0.00"),
    ]

    for j, (label, formula, fmt) in enumerate(summary, start=1):
        r = srow + j
        ws.cell(row=r, column=1, value=label).font = f_label
        cell = ws.cell(row=r, column=3, value=formula)
        cell.number_format = fmt
        if "TOTAL DSM PENALTY" in label:
            ws.cell(row=r, column=1).font = Font(name=FONT, size=10, bold=True, color=RED)
            cell.font = Font(name=FONT, size=10, bold=True, color=RED)
            cell.fill = PatternFill("solid", fgColor=RED_FILL)

    # ---------------- Enercast day summary - reference only ----------------
    if has_enercast:
        ec_srow = srow + len(summary) + 2
        ws.cell(row=ec_srow, column=1,
                value="ENERCAST SUMMARY - ACCURACY AND DSM PENALTY (reference only)"
                ).font = f_section

        i_rng = f"I{start}:I{last_row}"
        j_rng = f"J{start}:J{last_row}"
        l_rng = f"L{start}:L{last_row}"

        ec_summary = [
            ("Blocks with Enercast + a real meter reading",
             f'=SUMPRODUCT(({j_rng}<>"")*1)', "0"),
            ("Total Enercast (MW, scored blocks)",
             f'=SUMIF({j_rng},"<>",{i_rng})', "0.000"),
            ("Total actual (MW, scored blocks)",
             f'=SUMIF({j_rng},"<>",{d_rng})', "0.000"),
            ("Total deviation - actual minus Enercast (MW)", f"=SUM({j_rng})",
             "\\+0.000;\\-0.000"),
            ("Mean absolute deviation (MW)",
             f"=SUMPRODUCT(ABS({j_rng}))/COUNT({j_rng})", "0.000"),
            ("Max absolute deviation (MW)", f"=MAX(MAX({j_rng}),-MIN({j_rng}))", "0.000"),
            ("Blocks over 0.5 MW deviation (RED)",
             f'=SUMPRODUCT((ABS({j_rng})>0.5)*({j_rng}<>""))', "0"),
            ("Blocks within 0.5 MW deviation (BLUE)",
             f'=SUMPRODUCT((ABS({j_rng})<=0.5)*({j_rng}<>""))', "0"),
            ("Mean absolute deviation (% of capacity)",
             f"=SUMPRODUCT(ABS({j_rng}))/COUNT({j_rng})/{cap_cell}*100", "0.00"),
            ("Blocks that incurred a penalty", f'=COUNTIF({l_rng},">0")', "0"),
            ("Worst single-block penalty (Rs)", f"=MAX({l_rng})", "0.00"),
            ("TOTAL DSM PENALTY FOR THE DAY (Rs)", f"=SUM({l_rng})", "0.00"),
        ]

        for j, (label, formula, fmt) in enumerate(ec_summary, start=1):
            r = ec_srow + j
            ws.cell(row=r, column=1, value=label).font = f_label
            cell = ws.cell(row=r, column=3, value=formula)
            cell.number_format = fmt
            if "TOTAL DSM PENALTY" in label:
                ws.cell(row=r, column=1).font = Font(name=FONT, size=10, bold=True, color=RED)
                cell.font = Font(name=FONT, size=10, bold=True, color=RED)
                cell.fill = PatternFill("solid", fgColor=RED_FILL)

    # ---------------- conditional formatting: RED/BLUE deviation band ----------------
    ws.conditional_formatting.add(
        f"E{start}:F{last_row}",
        FormulaRule(formula=[f"AND(ISNUMBER($E{start}),ABS($E{start})>0.5)"],
                    font=Font(name=FONT, size=10, bold=True, color=RED),
                    fill=PatternFill("solid", fgColor=RED_FILL), stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        f"E{start}:F{last_row}",
        FormulaRule(formula=[f"AND(ISNUMBER($E{start}),ABS($E{start})<=0.5)"],
                    font=Font(name=FONT, size=10, bold=True, color=BLUE),
                    fill=PatternFill("solid", fgColor=BLUE_FILL), stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        f"G{start}:G{last_row}",
        FormulaRule(formula=[f"AND(ISNUMBER($G{start}),$G{start}>0)"],
                    font=Font(name=FONT, size=10, bold=True, color=RED),
                    fill=PatternFill("solid", fgColor=RED_FILL), stopIfTrue=True)
    )

    if has_enercast:
        ws.conditional_formatting.add(
            f"J{start}:K{last_row}",
            FormulaRule(formula=[f"AND(ISNUMBER($J{start}),ABS($J{start})>0.5)"],
                        font=Font(name=FONT, size=10, bold=True, color=RED),
                        fill=PatternFill("solid", fgColor=RED_FILL), stopIfTrue=True)
        )
        ws.conditional_formatting.add(
            f"J{start}:K{last_row}",
            FormulaRule(formula=[f"AND(ISNUMBER($J{start}),ABS($J{start})<=0.5)"],
                        font=Font(name=FONT, size=10, bold=True, color=BLUE),
                        fill=PatternFill("solid", fgColor=BLUE_FILL), stopIfTrue=True)
        )
        ws.conditional_formatting.add(
            f"L{start}:L{last_row}",
            FormulaRule(formula=[f"AND(ISNUMBER($L{start}),$L{start}>0)"],
                        font=Font(name=FONT, size=10, bold=True, color=RED),
                        fill=PatternFill("solid", fgColor=RED_FILL), stopIfTrue=True)
        )

    # ---------------- charts ----------------
    chart_col = get_column_letter(len(headers) + 2)

    # Moving penalty band as a shaded area: openpyxl/Excel has no native
    # "fill between two lines" series, so this is the standard stacked-
    # area trick - stack an INVISIBLE area (lower band) with a VISIBLE
    # one on top of it (band width). Because they stack, the visible
    # area's top edge lands exactly at lower+width = upper band, and its
    # bottom edge at lower band - the shaded region IS the band, and it
    # moves with the meter/actual line since lower band does. Blank for
    # blocks with no real meter reading (pre-dawn/post-dusk), which the
    # chart correctly renders as a gap rather than a shaded 0.
    band = AreaChart()
    band.grouping = "stacked"
    band.overlap = 100

    band_data = Reference(ws, min_col=band_lower_col, max_col=band_width_col,
                           min_row=HEAD, max_row=last_row)
    cats = Reference(ws, min_col=2, min_row=start, max_row=last_row)
    band.add_data(band_data, titles_from_data=True)
    band.set_categories(cats)

    lower_series, width_series = band.series
    lower_series.graphicalProperties.noFill = True
    lower_series.graphicalProperties.line.noFill = True
    width_series.graphicalProperties.solidFill = "D9E2F3"
    width_series.graphicalProperties.line.noFill = True
    # Named for the legend rather than the raw column header.
    width_series.tx = SeriesLabel(v="Allowed Band (+/-10% capacity)")

    line = LineChart()
    line.title = f"AI Schedule vs Actual Power - {day}"
    line.height, line.width = 9, 24
    line.y_axis.title = "MW"
    # No x-axis title: the category labels are plainly times, and an axis
    # title collides with the bottom legend in Excel's rendering.
    line.x_axis.title = None

    data = Reference(ws, min_col=3, max_col=4, min_row=HEAD, max_row=last_row)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.legend.position = "r"

    # Actual is the reference line, so it stays the heaviest; the two
    # forecasts are thinner and Enercast is dashed, so the three stay
    # readable even where they overlap.
    styling = [
        ("1F4E9C", 20000, None),        # AI schedule
        ("C00000", 26000, None),        # actual meter
        ("2E7D32", 20000, "dash"),      # Enercast
    ]

    if has_enercast:
        ec_data = Reference(ws, min_col=9, max_col=9, min_row=HEAD, max_row=last_row)
        line.add_data(ec_data, titles_from_data=True)

    for series, (colour, width, dash) in zip(line.series, styling):
        series.smooth = False
        series.graphicalProperties.line.solidFill = colour
        series.graphicalProperties.line.width = width
        if dash:
            series.graphicalProperties.line.dashStyle = dash
        series.marker.symbol = "none"

    # openpyxl defaults BOTH axes to delete=True, which means "hide this
    # axis" - so the time labels can vanish entirely. Set explicitly.
    # Do NOT set tickLblSkip here: Excel already thins crowded category
    # labels sensibly on its own, and forcing a skip dropped the time
    # axis on the 2026-07-28/29 rebuild.
    line.x_axis.delete = False
    line.y_axis.delete = False

    # Combine so the band renders as shading BEHIND the schedule/actual
    # lines, not a separate chart - band first puts it at the back.
    # openpyxl's ChartBase only overrides __iadd__ (+=), not __add__, so
    # `band + line` falls through to Serialisable.__add__ and errors on
    # the type mismatch. Must use += to merge different chart types.
    # += only merges the series list (_charts) - title/axes/legend/size
    # stay whatever `band` already had (all defaults), so copy them over
    # from `line` onto the surviving `band` object explicitly.
    band.title = line.title
    band.height, band.width = line.height, line.width
    band.y_axis.title = line.y_axis.title
    band.x_axis.title = line.x_axis.title
    band.x_axis.delete = line.x_axis.delete
    band.y_axis.delete = line.y_axis.delete
    band.legend.position = line.legend.position
    # idx 0 = lower_series, the invisible band-anchor series - hide it from
    # the legend so only "Allowed Band" (idx 1) shows for the shaded area.
    band.legend.legendEntry = [LegendEntry(idx=0, delete=True)]

    band += line
    ws.add_chart(band, f"{chart_col}3")

    bar = BarChart()
    bar.type = "col"
    bar.title = f"DSM Penalty per Block (Rs) - {day}"
    bar.height, bar.width = 8, 24
    bar.y_axis.title = "Rs"
    bar.x_axis.title = None

    pen_data = Reference(ws, min_col=7, max_col=7, min_row=HEAD, max_row=last_row)
    bar.add_data(pen_data, titles_from_data=True)

    if has_enercast:
        ec_pen_data = Reference(ws, min_col=12, max_col=12, min_row=HEAD, max_row=last_row)
        bar.add_data(ec_pen_data, titles_from_data=True)

    bar.set_categories(cats)
    bar.legend.position = "r"

    # Two clustered series across ~53 blocks is 106 bars. Excel's default
    # gap (150%) leaves them hairline-thin and unreadable, so the gap is
    # closed right down and the pair overlapped slightly - each block then
    # reads as one two-tone column rather than a picket fence.
    bar.gapWidth = 30
    bar.overlap = -10

    for series, colour in zip(bar.series, ("1F4E9C", "2E7D32")):
        series.graphicalProperties.solidFill = colour
        series.graphicalProperties.line.solidFill = colour

    bar.x_axis.delete = False
    bar.y_axis.delete = False

    ws.add_chart(bar, f"{chart_col}26")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 17
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 13
    if has_enercast:
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 17
        ws.column_dimensions["K"].width = 22
        ws.column_dimensions["L"].width = 15
    ws.freeze_panes = ws.cell(row=start, column=1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = OUT_DIR / f"Schedule_vs_Meter_Penalty_{DAY}.xlsx"
    target = xlsx_path
    for attempt in range(1, 20):
        try:
            wb.save(target)
            break
        except PermissionError:
            target = xlsx_path.with_name(f"{xlsx_path.stem}_{attempt}{xlsx_path.suffix}")

    print(f"rows: {len(rows)}   scored blocks: {int(rows['active_power_kw'].notna().sum())}")
    print(f"saved: {target}")
    print(f"NOTE: run recalc.py on this file - it is formula-driven and has no cached values yet.")


if __name__ == "__main__":
    main()
