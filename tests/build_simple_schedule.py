"""
=========================================================
Solar Forecasting Project
Simple Schedule vs Meter Comparison
=========================================================
The detailed workbook (one column per scheduling time) was
too dense to read against the meter file, so this emits the
plainest possible comparison, laid out exactly like the
meter CSV the mentor already reads:

    TimeStamp, Active Power (kW), ...

One row per 15-minute block, the SAME timestamps as the
meter file, and one power value per row. Nothing is
accumulated - each row is that block's own power, exactly
as the meter records it. Rows the meter covers but the
schedule does not (before the first scheduling run) are
left blank rather than dropped, so the two files line up
row for row and can be pasted side by side.

Run:  python -m tests.build_simple_schedule [YYYY-MM-DD]
=========================================================
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.config import settings
from modules.preprocessing.preprocess import DataPreprocessor
from modules.evaluation.evaluator import Evaluator


DAY = sys.argv[1] if len(sys.argv) > 1 else "2026-07-25"
SRC = Path("outputs/schedules")
OUT_DIR = Path("outputs/reports")

CAPACITY_KW = settings["plant"]["capacity_mw"] * 1000
PLANT = settings["plant"]["name"]

FONT = "Arial"
NAVY = "1F3864"
GREY = "F2F2F2"


def load_actual(day):
    """The day's meter readings, straight from the raw file."""

    path = Path(settings["paths"]["historical_data"]) / f"{day.replace('-', '_')}_SOLAR_INV.csv"

    frame = DataPreprocessor().preprocess(
        file_path=path,
        required_columns=["TimeStamp"],
        timestamp_column="TimeStamp",
    )

    columns = ["timestamp", "active_power_kw"]
    if "is_real_measurement" in frame.columns:
        columns.append("is_real_measurement")

    return frame[columns]


def main():

    schedule = pd.read_csv(SRC / f"day_schedule_{DAY}.csv", parse_dates=["timestamp"])
    actual = load_actual(DAY)

    # Every timestamp either side knows about, so the file lines up with
    # the meter CSV row for row.
    merged = actual.merge(
        schedule[["timestamp", "scheduled_mw"]], on="timestamp", how="outer"
    ).sort_values("timestamp").reset_index(drop=True)

    merged = merged[merged["timestamp"].dt.date == pd.Timestamp(DAY).date()]

    # Drop the pre-dawn blocks the meter covers but no scheduling run ever
    # reached - the first run is at 06:45, so nothing before 07:00 has a
    # schedule to compare against and those rows were only empty space.
    merged = merged[merged["scheduled_mw"].notna()].reset_index(drop=True)

    # Enercast, side by side, for comparison only - never used to grade
    # our own forecast (see modules/evaluation/evaluator.py). Degrades
    # gracefully: a day with no Enercast file just gets no extra columns.
    enercast_day = Evaluator().load_enercast_day(pd.Timestamp(DAY))
    has_enercast = enercast_day is not None
    if has_enercast:
        merged = merged.merge(enercast_day, on="timestamp", how="left")

    # MW throughout - the plant is rated in MW and mixed units caused
    # repeated confusion. Per-block deviation sits beside the error so
    # each block's own accuracy is visible, not just the day average.
    capacity_mw = CAPACITY_KW / 1000

    out = pd.DataFrame({
        "Time": merged["timestamp"].dt.strftime("%H:%M"),
        "Scheduled (MW)": merged["scheduled_mw"].round(4),
        "Actual (MW)": (merged["active_power_kw"] / 1000).round(4),
    })
    out["Error (MW)"] = (
        out["Scheduled (MW)"] - out["Actual (MW)"]
    ).round(4)
    out["Deviation (%)"] = (
        out["Error (MW)"].abs() / capacity_mw * 100
    ).round(2)

    if has_enercast:
        out["Enercast (MW)"] = (merged["enercast_kw"] / 1000).round(4)
        out["Enercast Error (MW)"] = (
            out["Enercast (MW)"] - out["Actual (MW)"]
        ).round(4)
        out["Enercast Deviation (%)"] = (
            out["Enercast Error (MW)"].abs() / capacity_mw * 100
        ).round(2)

    csv_path = OUT_DIR / f"Schedule_vs_Meter_{DAY}.csv"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out.to_csv(csv_path, index=False)

    # ---------------- the same thing as a workbook ----------------
    wb = Workbook()
    ws = wb.active
    ws.title = f"Schedule vs Meter"

    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")
    f_head = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    f_body = Font(name=FONT, size=10)
    f_bold = Font(name=FONT, size=10, bold=True)
    fill_head = PatternFill("solid", fgColor=NAVY)
    fill_grey = PatternFill("solid", fgColor=GREY)

    ws.cell(row=1, column=1,
            value=f"{PLANT} - Scheduled vs Actual - {DAY}  "
                  f"(15-minute blocks, all values in MW)"
            ).font = Font(name=FONT, size=12, bold=True, color=NAVY)

    if has_enercast:
        ws.cell(row=2, column=1,
                value="Enercast is shown alongside for comparison only - our forecast is "
                      "graded against actual meter data, never against Enercast."
                ).font = Font(name=FONT, size=9, italic=True, color="595959")

    headers = list(out.columns)
    formats = {
        "Scheduled (MW)": "0.0000",
        "Actual (MW)": "0.0000",
        "Error (MW)": "0.0000;-0.0000",
        "Deviation (%)": "0.00",
        "Enercast (MW)": "0.0000",
        "Enercast Error (MW)": "0.0000;-0.0000",
        "Enercast Deviation (%)": "0.00",
    }

    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = f_head
        cell.fill = fill_head
        cell.alignment = center
        cell.border = box

    for r, (_, rec) in enumerate(out.iterrows(), start=4):
        for i, key in enumerate(headers, start=1):
            value = rec[key]
            if pd.isna(value):
                value = None
            cell = ws.cell(row=r, column=i, value=value)
            cell.font = f_body
            cell.border = box
            cell.alignment = center
            if key in formats:
                cell.number_format = formats[key]

    last = 3 + len(out)

    # ---------------- accuracy, in MW only ----------------
    # Blocks the meter actually measured; the trailing blocks after 18:00
    # are scheduled but have no reading yet, so they cannot be scored.
    both = out.dropna(subset=["Scheduled (MW)", "Actual (MW)"])

    scheduled_mw = both["Scheduled (MW)"]
    actual_mw = both["Actual (MW)"]
    error_mw = both["Error (MW)"]

    total_predicted = float(scheduled_mw.sum())
    total_actual = float(actual_mw.sum())
    total_error = total_predicted - total_actual
    total_abs_error = float(error_mw.abs().sum())
    mae = float(error_mw.abs().mean())
    rmse = float((error_mw ** 2).mean() ** 0.5)
    deviation = mae / capacity_mw * 100
    worst = both.loc[both["Deviation (%)"].idxmax()]

    mrow = last + 2
    ws.cell(row=mrow, column=1, value="ACCURACY vs ACTUAL METER DATA").font = Font(
        name=FONT, size=11, bold=True, color=NAVY)

    # Energy figures are deliberately absent: they can only be stated in
    # MWh, and this report is MW-only by request.
    metrics = [
        ("Blocks scored (real meter readings)", len(both), "0"),
        ("Average predicted (MW)", round(float(scheduled_mw.mean()), 4), "0.0000"),
        ("Average actual (MW)", round(float(actual_mw.mean()), 4), "0.0000"),
        ("Peak predicted (MW)", round(float(scheduled_mw.max()), 4), "0.0000"),
        ("Peak actual (MW)", round(float(actual_mw.max()), 4), "0.0000"),
        ("Total predicted (MW)", round(total_predicted, 4), "0.0000"),
        ("Total actual (MW)", round(total_actual, 4), "0.0000"),
        ("Total error - predicted minus actual (MW)", round(total_error, 4),
         "+0.0000;-0.0000"),
        ("Total absolute error (MW)", round(total_abs_error, 4), "0.0000"),
        ("Average percentage deviation (%)", round(deviation, 2), "0.00"),
        ("Worst block deviation (%)", round(float(worst["Deviation (%)"]), 2), "0.00"),
        ("Worst block time", str(worst["Time"]), None),
        ("Mean absolute error (MW)", round(mae, 4), "0.0000"),
        ("Root mean squared error (MW)", round(rmse, 4), "0.0000"),
    ]

    for j, (label, value, fmt) in enumerate(metrics, start=1):
        ws.cell(row=mrow + j, column=1, value=label).font = f_bold
        cell = ws.cell(row=mrow + j, column=3, value=value)
        cell.font = f_body
        if fmt:
            cell.number_format = fmt
        cell.fill = fill_grey
        cell.border = box
        cell.alignment = center

    erow = mrow + len(metrics)

    # ---------------- Enercast, side by side - reference only ----------------
    # Same metrics, computed against the same scored blocks, so the two
    # forecasts are judged on identical footing. Never used to grade our
    # own forecast - see modules/evaluation/evaluator.py.
    if has_enercast:
        both_ec = out.dropna(subset=["Enercast (MW)", "Actual (MW)"])

        if not both_ec.empty:
            ec_mw = both_ec["Enercast (MW)"]
            ec_actual_mw = both_ec["Actual (MW)"]
            ec_error_mw = both_ec["Enercast Error (MW)"]

            ec_total_predicted = float(ec_mw.sum())
            ec_total_actual = float(ec_actual_mw.sum())
            ec_total_error = ec_total_predicted - ec_total_actual
            ec_total_abs_error = float(ec_error_mw.abs().sum())
            ec_mae = float(ec_error_mw.abs().mean())
            ec_rmse = float((ec_error_mw ** 2).mean() ** 0.5)
            ec_deviation = ec_mae / capacity_mw * 100
            ec_worst = both_ec.loc[both_ec["Enercast Deviation (%)"].idxmax()]

            erow += 2
            ws.cell(row=erow, column=1,
                    value="ENERCAST vs ACTUAL METER DATA (reference only)"
                    ).font = Font(name=FONT, size=11, bold=True, color=NAVY)

            enercast_metrics = [
                ("Blocks scored (real meter readings)", len(both_ec), "0"),
                ("Average Enercast (MW)", round(float(ec_mw.mean()), 4), "0.0000"),
                ("Average actual (MW)", round(float(ec_actual_mw.mean()), 4), "0.0000"),
                ("Peak Enercast (MW)", round(float(ec_mw.max()), 4), "0.0000"),
                ("Peak actual (MW)", round(float(ec_actual_mw.max()), 4), "0.0000"),
                ("Total Enercast (MW)", round(ec_total_predicted, 4), "0.0000"),
                ("Total actual (MW)", round(ec_total_actual, 4), "0.0000"),
                ("Total error - Enercast minus actual (MW)", round(ec_total_error, 4),
                 "+0.0000;-0.0000"),
                ("Total absolute error (MW)", round(ec_total_abs_error, 4), "0.0000"),
                ("Average percentage deviation (%)", round(ec_deviation, 2), "0.00"),
                ("Worst block deviation (%)", round(float(ec_worst["Enercast Deviation (%)"]), 2),
                 "0.00"),
                ("Worst block time", str(ec_worst["Time"]), None),
                ("Mean absolute error (MW)", round(ec_mae, 4), "0.0000"),
                ("Root mean squared error (MW)", round(ec_rmse, 4), "0.0000"),
            ]

            for j, (label, value, fmt) in enumerate(enercast_metrics, start=1):
                ws.cell(row=erow + j, column=1, value=label).font = f_bold
                cell = ws.cell(row=erow + j, column=3, value=value)
                cell.font = f_body
                if fmt:
                    cell.number_format = fmt
                cell.fill = fill_grey
                cell.border = box
                cell.alignment = center

            print()
            print("ACCURACY - ENERCAST (reference only)")
            for label, value, _fmt in enercast_metrics:
                print(f"  {label:<44} {value}")

    # ---------------- scheduled vs actual, as a line chart ----------------
    # The morning-under / afternoon-over pattern is obvious in a chart and
    # invisible in a column of numbers.
    chart = LineChart()
    chart.title = f"Scheduled vs Actual - {DAY}"
    chart.style = 2
    chart.height = 9
    chart.width = 24
    chart.y_axis.title = "MW"
    chart.x_axis.title = "Time"
    chart.y_axis.majorGridlines = ChartLines()

    data = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=last)
    times = Reference(ws, min_col=1, min_row=4, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(times)

    colours = ["1F4E79", "C00000"]

    if has_enercast:
        # Enercast column is not adjacent to Scheduled/Actual (Error and
        # Deviation sit between them), so it is its own Reference rather
        # than one contiguous range.
        enercast_col = headers.index("Enercast (MW)") + 1
        ec_data = Reference(ws, min_col=enercast_col, max_col=enercast_col,
                             min_row=3, max_row=last)
        chart.add_data(ec_data, titles_from_data=True)
        colours.append("2E7D32")

    for series, colour in zip(chart.series, colours):
        series.smooth = False
        series.graphicalProperties.line.solidFill = colour
        series.graphicalProperties.line.width = 22000   # EMU, ~1.75pt
        series.marker.symbol = "none"

    chart_anchor_col = get_column_letter(len(headers) + 2)
    ws.add_chart(chart, f"{chart_anchor_col}3")

    ws.column_dimensions["A"].width = 34
    for col in headers[1:]:
        ws.column_dimensions[get_column_letter(headers.index(col) + 1)].width = 17
    ws.freeze_panes = ws.cell(row=4, column=1)

    xlsx_path = OUT_DIR / f"Schedule_vs_Meter_{DAY}.xlsx"
    target = xlsx_path
    for attempt in range(1, 20):
        try:
            wb.save(target)
            break
        except PermissionError:
            target = xlsx_path.with_name(
                f"{xlsx_path.stem}_{attempt}{xlsx_path.suffix}")

    print(f"rows: {len(out)}   scored blocks: {len(both)}")
    print()
    print(out.head(6).to_string(index=False))
    print("...")
    print(out.tail(4).to_string(index=False))
    print()
    print("ACCURACY (MW only)")
    for label, value, _fmt in metrics:
        print(f"  {label:<44} {value}")
    print()
    print(f"  line chart embedded at cell {chart_anchor_col}3")
    print()
    print(f"saved CSV  : {csv_path}")
    print(f"saved Excel: {target}")


if __name__ == "__main__":
    main()
