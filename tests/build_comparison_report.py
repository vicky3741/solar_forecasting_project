"""
=========================================================
Solar Forecasting Project
Our Schedule vs Friend's Schedule vs Actual + Enercast
=========================================================
Fast side-by-side comparison for a day where a teammate ran
their own (modified) pipeline and produced their own AI
schedule. Actual meter and Enercast come from OUR already-
built Schedule_vs_Meter_Penalty report (they are identical in
both files - neither pipeline touches that data). Only the
AI Schedule column differs between the two pipelines, so this
report pulls:

  - Block / Time / Actual / Enercast (+ Enercast deviation,
    deviation %, penalty) straight from our own report - not
    recomputed, so Enercast stays exactly as it already was.
  - Our AI Schedule from our own report.
  - Friend's AI Schedule from their delivered xlsx.

Our deviation/penalty is a live formula against the same DSM
slab parameters build_penalty_report.py uses. Friend's penalty
is copied as-is from their file (their pipeline change may
compute it differently, so recomputing it here would misrepresent
their result) - flagged clearly as "as reported".

Run:  python -m tests.build_comparison_report [YYYY-MM-DD] [friend_xlsx_path]
=========================================================
"""

import argparse
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.config import settings

def parse_args():
    """
    Positionals keep the original `... <day> [friend_xlsx]` form working.
    --ours-xlsx exists because build_penalty_report.py falls back to a
    `_1` filename when the canonical report is open in Excel; without it
    this script would silently read the stale locked copy and pair a fresh
    friend column with an out-of-date Enercast one.
    """

    parser = argparse.ArgumentParser()
    parser.add_argument("day", nargs="?", default="2026-08-06")
    parser.add_argument(
        "friend_xlsx", nargs="?",
        default=r"C:\Users\Acer\Downloads\Team2_Sirmour_Schedule_2026-08-06_Final.xlsx",
    )
    parser.add_argument(
        "--ours-xlsx", dest="ours_xlsx", default=None,
        help="our own report to read Actual/Enercast/our schedule from "
             "(default: outputs/reports/Schedule_vs_Meter_Penalty_<day>.xlsx)",
    )
    return parser.parse_args()


ARGS = parse_args()
DAY = ARGS.day
FRIEND_PATH = Path(ARGS.friend_xlsx)
OUT_DIR = Path("outputs/reports")
OURS_PATH = Path(ARGS.ours_xlsx) if ARGS.ours_xlsx else (
    OUT_DIR / f"Schedule_vs_Meter_Penalty_{DAY}.xlsx"
)

CAPACITY_MW = settings["plant"]["capacity_mw"]
PLANT = settings["plant"]["name"]
BLOCK_ENERGY_FACTOR = 250

FONT = "Arial"
DARK_GREEN = "1B4332"
RED = "C00000"
RED_FILL = "FCE4E4"
BLUE = "1F4E9C"
PURPLE = "6A3D9A"

SLABS = [
    ("Slab 1", 0, 10, 0),
    ("Slab 2", 10, 15, 0.5),
    ("Slab 3", 15, 20, 0.75),
    ("Slab 4", 20, None, 1),
]


def load_ours(path):
    """Block -> dict of our own report's per-block values (cached, since
    build_penalty_report.py always leaves the file recalculated)."""

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        if row[0].value == "Block":
            header_row = row[0].row
            break

    rows = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        block = row[0].value
        if not isinstance(block, int):
            break
        rows[block] = {
            "time": row[1].value,
            "our_schedule": row[2].value,
            "actual": row[3].value,
            "enercast": row[8].value,
            "enercast_dev": row[9].value,
            "enercast_dev_pct": row[10].value,
            "enercast_penalty": row[11].value,
        }
    return rows


def load_friend(path):
    """Block -> friend's AI Schedule + their own reported deviation/penalty."""

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        if row[0].value == "Block":
            header_row = row[0].row
            break

    rows = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        block = row[0].value
        if not isinstance(block, int):
            break
        rows[block] = {
            "friend_schedule": row[2].value,
            "friend_penalty": row[6].value,
        }
    return rows


def main():

    ours = load_ours(OURS_PATH)
    friend = load_friend(FRIEND_PATH)

    blocks = sorted(set(ours) & set(friend))
    missing_ours = sorted(set(friend) - set(ours))
    missing_friend = sorted(set(ours) - set(friend))

    wb = Workbook()
    ws = wb.active
    ws.title = "Ours vs Friend vs Actual"

    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")
    f_head = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    f_body = Font(name=FONT, size=10)
    f_section = Font(name=FONT, size=12, bold=True, color=DARK_GREEN)
    f_label = Font(name=FONT, size=10)
    f_input = Font(name=FONT, size=10, color="0000FF")
    f_note = Font(name=FONT, size=9, italic=True, color="595959")
    fill_head = PatternFill("solid", fgColor=DARK_GREEN)

    headers = [
        "Block", "Time", "Our AI Schedule (MW)", "Friend AI Schedule (MW)",
        "Actual (MW)", "Our Deviation (MW)", "Friend Deviation (MW)",
        "Our Deviation % (Capacity)", "Friend Deviation % (Capacity)",
        "Our Penalty (Rs)", "Friend Penalty (Rs, as reported)",
        "Enercast (MW)", "Enercast Deviation (MW)",
        "Enercast Deviation % (Capacity)", "Enercast Penalty (Rs)",
    ]
    last_col_letter = get_column_letter(len(headers))

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws.cell(row=1, column=1,
            value=f"{PLANT} - Our Schedule vs Friend's Schedule vs Actual - {DAY}"
            ).font = f_section
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws.cell(row=2, column=1,
            value="Actual meter + Enercast are identical in both pipelines' outputs, taken "
                  "once from our own report. Only the AI Schedule differs by pipeline. "
                  "Friend's penalty is copied as-is (their pipeline may compute it "
                  "differently) - Our Penalty is a live formula against the DSM slabs below."
            ).font = f_note

    HEAD = 4
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=HEAD, column=i, value=h)
        cell.font = f_head
        cell.fill = fill_head
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box

    start = HEAD + 1

    for r, block in enumerate(blocks, start=start):
        o = ours[block]
        fr = friend[block]

        ws.cell(row=r, column=1, value=block)
        ws.cell(row=r, column=2, value=o["time"])

        ws.cell(row=r, column=3, value=o["our_schedule"]).number_format = "0.000"
        ws.cell(row=r, column=4, value=fr["friend_schedule"]).number_format = "0.000"

        has_actual = o["actual"] is not None
        if has_actual:
            ws.cell(row=r, column=5, value=o["actual"]).number_format = "0.000"
            ws.cell(row=r, column=6, value=f"=E{r}-C{r}").number_format = "0.000"
            ws.cell(row=r, column=7, value=f"=E{r}-D{r}").number_format = "0.000"

        if o["enercast"] is not None:
            ws.cell(row=r, column=12, value=o["enercast"]).number_format = "0.000"
        if o["enercast_dev"] is not None:
            ws.cell(row=r, column=13, value=o["enercast_dev"]).number_format = "0.000"
        if o["enercast_dev_pct"] is not None:
            ws.cell(row=r, column=14, value=o["enercast_dev_pct"]).number_format = "\\+0.00;\\-0.00"
        if o["enercast_penalty"] is not None:
            ws.cell(row=r, column=15, value=o["enercast_penalty"]).number_format = "0.00"

        if fr["friend_penalty"] is not None:
            ws.cell(row=r, column=11, value=fr["friend_penalty"]).number_format = "0.00"

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = box
            cell.alignment = center
            if cell.font is None or cell.font.sz is None:
                cell.font = f_body

    last_row = start + len(blocks) - 1

    # ---------------- DSM slab parameters ----------------
    param_row = last_row + 3
    cap_cell = f"C{param_row + 1}"
    factor_cell = f"C{param_row + 2}"

    ws.cell(row=param_row, column=1,
            value="DSM SLAB PARAMETERS (Our Deviation %/Penalty reference these cells)"
            ).font = f_section
    ws.cell(row=param_row + 1, column=1, value="Installed capacity (MW)").font = f_label
    ws.cell(row=param_row + 1, column=3, value=CAPACITY_MW).font = f_input
    ws.cell(row=param_row + 2, column=1,
            value="Block energy factor (0.25 h x 1000 kW/MW)").font = f_label
    ws.cell(row=param_row + 2, column=3, value=BLOCK_ENERGY_FACTOR).font = f_input

    slab_head_row = param_row + 4
    for i, h in enumerate(["Slab", "From %", "To %", "Rate (Rs/kWh)", "Upper edge (MW)"], start=1):
        cell = ws.cell(row=slab_head_row, column=i, value=h)
        cell.font = f_head
        cell.fill = fill_head
        cell.alignment = center

    slab_rows = {}
    for i, (label, frm, to, rate) in enumerate(SLABS, start=1):
        r = slab_head_row + i
        slab_rows[label] = r
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=frm)
        ws.cell(row=r, column=3, value=to if to is not None else "above")
        ws.cell(row=r, column=4, value=rate).font = f_input
        if to is not None:
            ws.cell(row=r, column=5, value=f"={cap_cell}*{to}/100").number_format = "0.000"

    slab1_edge = f"E{slab_rows['Slab 1']}"
    slab2_edge = f"E{slab_rows['Slab 2']}"
    slab3_edge = f"E{slab_rows['Slab 3']}"
    slab2_rate = f"D{slab_rows['Slab 2']}"
    slab3_rate = f"D{slab_rows['Slab 3']}"
    slab4_rate = f"D{slab_rows['Slab 4']}"

    def penalty_formula(dev_col, r):
        return (
            f"={factor_cell}*("
            f"MAX(0,MIN(ABS({dev_col}{r}),{slab2_edge})-{slab1_edge})*{slab2_rate}"
            f"+MAX(0,MIN(ABS({dev_col}{r}),{slab3_edge})-{slab2_edge})*{slab3_rate}"
            f"+MAX(0,ABS({dev_col}{r})-{slab3_edge})*{slab4_rate})"
        )

    for r in range(start, last_row + 1):
        if ws.cell(row=r, column=5).value is None:
            continue
        pct_cell = ws.cell(row=r, column=8, value=f"=F{r}/{cap_cell}*100")
        pct_cell.number_format = "\\+0.00;\\-0.00"
        pen_cell = ws.cell(row=r, column=10, value=penalty_formula("F", r))
        pen_cell.number_format = "0.00"

        friend_pct_cell = ws.cell(row=r, column=9, value=f"=G{r}/{cap_cell}*100")
        friend_pct_cell.number_format = "\\+0.00;\\-0.00"

    # ---------------- day summary ----------------
    srow = last_row + 2 + 8  # leave slab block above, summary further down
    srow = slab_head_row + len(SLABS) + 3

    ws.cell(row=srow, column=1, value="DAY SUMMARY - OURS vs FRIEND vs ENERCAST").font = f_section

    f_rng = f"F{start}:F{last_row}"
    g_rng = f"G{start}:G{last_row}"
    j_rng = f"J{start}:J{last_row}"
    k_rng = f"K{start}:K{last_row}"
    m_rng = f"M{start}:M{last_row}"
    o_rng = f"O{start}:O{last_row}"

    summary = [
        ("Total our penalty (Rs)", f"=SUM({j_rng})", "0.00"),
        ("Total friend penalty (Rs, as reported)", f"=SUM({k_rng})", "0.00"),
        ("Total Enercast penalty (Rs, reference)", f"=SUM({o_rng})", "0.00"),
        ("Mean abs our deviation (MW)", f"=SUMPRODUCT(ABS({f_rng}))/COUNT({f_rng})", "0.000"),
        ("Mean abs friend deviation (MW)", f"=SUMPRODUCT(ABS({g_rng}))/COUNT({g_rng})", "0.000"),
        ("Mean abs Enercast deviation (MW)", f"=SUMPRODUCT(ABS({m_rng}))/COUNT({m_rng})", "0.000"),
    ]
    for j, (label, formula, fmt) in enumerate(summary, start=1):
        r = srow + j
        ws.cell(row=r, column=1, value=label).font = f_label
        cell = ws.cell(row=r, column=3, value=formula)
        cell.number_format = fmt
        if "penalty" in label.lower() and "total" in label.lower():
            ws.cell(row=r, column=1).font = Font(name=FONT, size=10, bold=True, color=RED)
            cell.font = Font(name=FONT, size=10, bold=True, color=RED)
            cell.fill = PatternFill("solid", fgColor=RED_FILL)

    if missing_ours or missing_friend:
        note_row = srow + len(summary) + 2
        note = []
        if missing_ours:
            note.append(f"blocks only in friend's file (no match in ours): {missing_ours}")
        if missing_friend:
            note.append(f"blocks only in ours (no match in friend's file): {missing_friend}")
        ws.cell(row=note_row, column=1, value="NOTE: " + "; ".join(note)).font = f_note

    # ---------------- conditional formatting ----------------
    ws.conditional_formatting.add(
        f"F{start}:F{last_row}",
        FormulaRule(formula=[f"AND(ISNUMBER($F{start}),ABS($F{start})>0.5)"],
                    font=Font(name=FONT, size=10, bold=True, color=RED),
                    fill=PatternFill("solid", fgColor=RED_FILL), stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        f"G{start}:G{last_row}",
        FormulaRule(formula=[f"AND(ISNUMBER($G{start}),ABS($G{start})>0.5)"],
                    font=Font(name=FONT, size=10, bold=True, color=PURPLE),
                    fill=PatternFill("solid", fgColor="F1E4FA"), stopIfTrue=True)
    )

    # ---------------- chart ----------------
    chart_col = get_column_letter(len(headers) + 2)
    cats = Reference(ws, min_col=2, min_row=start, max_row=last_row)

    line = LineChart()
    line.title = f"Our Schedule vs Friend's Schedule vs Actual vs Enercast - {DAY}"
    line.height, line.width = 10, 26
    line.y_axis.title = "MW"
    line.x_axis.title = None

    data = Reference(ws, min_col=3, max_col=5, min_row=HEAD, max_row=last_row)
    line.add_data(data, titles_from_data=True)
    ec_data = Reference(ws, min_col=12, max_col=12, min_row=HEAD, max_row=last_row)
    line.add_data(ec_data, titles_from_data=True)
    line.set_categories(cats)
    line.legend.position = "r"

    styling = [
        (BLUE, 20000, None),      # our schedule
        (PURPLE, 20000, None),    # friend schedule
        (RED, 26000, None),       # actual
        ("2E7D32", 18000, "dash"),  # enercast
    ]
    for series, (colour, width, dash) in zip(line.series, styling):
        series.smooth = False
        series.graphicalProperties.line.solidFill = colour
        series.graphicalProperties.line.width = width
        if dash:
            series.graphicalProperties.line.dashStyle = dash
        series.marker.symbol = "none"

    line.x_axis.delete = False
    line.y_axis.delete = False
    ws.add_chart(line, f"{chart_col}3")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 15
    for col in "CDEFGHIJKLMNO":
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = ws.cell(row=start, column=1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / f"Schedule_Comparison_{DAY}.xlsx"
    target = out_path
    for attempt in range(1, 20):
        try:
            wb.save(target)
            break
        except PermissionError:
            target = out_path.with_name(f"{out_path.stem}_{attempt}{out_path.suffix}")

    print(f"blocks compared: {len(blocks)}")
    if missing_ours:
        print(f"blocks only in friend's file: {missing_ours}")
    if missing_friend:
        print(f"blocks only in our file: {missing_friend}")
    print(f"saved: {target}")


if __name__ == "__main__":
    main()
