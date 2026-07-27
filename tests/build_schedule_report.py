"""
Build the mentor-facing report workbook for one reconstructed day,
from the CSVs generate_schedule_for_day.py produces.

Sheets: Summary | Methodology | Day Schedule | Scheduling Runs
All power values in MW. Error metrics are live formulas so the sheet
recalculates if anyone edits it.
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DAY = sys.argv[1] if len(sys.argv) > 1 else "2026-07-25"
SRC = Path("outputs/schedules")
OUT = Path(f"outputs/reports/Schedule_Report_{DAY}.xlsx")

CAPACITY_MW = 5.1
PLANT = "Sirmour Solar Plant"
RUN_TIMES = "06:45, 08:15, 09:45, 11:15, 12:45, 14:15, 15:45"

FONT = "Arial"
NAVY = "1F3864"
LIGHT = "D9E2F3"
GREY = "F2F2F2"

schedule = pd.read_csv(SRC / f"day_schedule_{DAY}.csv")
run_log = pd.read_csv(SRC / f"day_schedule_{DAY}_run_log.csv")

thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, last_col):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box


def title_block(ws, title, subtitle, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.cell(row=1, column=1, value=title).font = Font(
        name=FONT, size=14, bold=True, color=NAVY)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws.cell(row=2, column=1, value=subtitle).font = Font(
        name=FONT, size=10, italic=True, color="595959")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22


wb = Workbook()

# ============================== SUMMARY ==============================
ws = wb.active
ws.title = "Summary"
title_block(ws, f"{PLANT} - Schedule Accuracy Report",
            f"Reconstructed day schedule for {DAY} | Plant capacity {CAPACITY_MW} MW", 4)

row = 4
ws.cell(row=row, column=1, value="PLANT & SCHEDULE DETAILS").font = Font(
    name=FONT, size=11, bold=True, color=NAVY)
row += 1

details = [
    ("Plant", PLANT),
    ("Installed capacity (MW)", CAPACITY_MW),
    ("Schedule date", DAY),
    ("Block duration", "15 minutes"),
    ("Scheduling times", RUN_TIMES),
    ("Windy clips available",
     f"{int((run_log['vision_signal'] == 'yes').sum())} of {len(run_log)}"),
    ("Blocks scheduled", len(schedule)),
]
for label, value in details:
    ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=row, column=2, value=value).font = Font(name=FONT, size=10)
    row += 1

row += 1
ws.cell(row=row, column=1, value="ACCURACY vs ACTUAL METER DATA").font = Font(
    name=FONT, size=11, bold=True, color=NAVY)
row += 1
ws.cell(row=row, column=1,
        value="Scored on real measured blocks only; interpolated blocks excluded."
        ).font = Font(name=FONT, size=9, italic=True, color="595959")
row += 1

metric_start = row
n = len(schedule)
# Live formulas against the Day Schedule sheet (rows 4..3+n there)
rng = lambda col: f"'Day Schedule'!{col}4:{col}{3 + n}"
flag = rng("G")          # actual_is_real
sched = rng("C")         # scheduled_mw
act = rng("D")           # actual_mw
err = rng("E")           # error_mw

metrics = [
    ("Blocks scored", f'=SUMPRODUCT(--({flag}=TRUE))', "0"),
    ("Average percentage deviation (%)",
     f'=SUMPRODUCT(--({flag}=TRUE),ABS({err}))/SUMPRODUCT(--({flag}=TRUE))/{CAPACITY_MW}*100',
     "0.00"),
    ("Mean absolute error (MW)",
     f'=SUMPRODUCT(--({flag}=TRUE),ABS({err}))/SUMPRODUCT(--({flag}=TRUE))', "0.0000"),
    ("Root mean squared error (MW)",
     f'=SQRT(SUMPRODUCT(--({flag}=TRUE),{err}^2)/SUMPRODUCT(--({flag}=TRUE)))', "0.0000"),
    ("Scheduled energy (MWh)",
     f'=SUMPRODUCT(--({flag}=TRUE),{sched})*0.25', "0.000"),
    ("Actual energy (MWh)",
     f'=SUMPRODUCT(--({flag}=TRUE),{act})*0.25', "0.000"),
    ("Energy difference (MWh)", f"=B{metric_start + 4}-B{metric_start + 5}", "+0.000;-0.000;0.000"),
    ("Average scheduled (MW)",
     f'=SUMPRODUCT(--({flag}=TRUE),{sched})/SUMPRODUCT(--({flag}=TRUE))', "0.000"),
    ("Average actual (MW)",
     f'=SUMPRODUCT(--({flag}=TRUE),{act})/SUMPRODUCT(--({flag}=TRUE))', "0.000"),
    ("Peak scheduled (MW)", f'=MAX({sched})', "0.000"),
    ("Peak actual (MW)", f'=MAX({act})', "0.000"),
]

for label, formula, fmt in metrics:
    ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True)
    c = ws.cell(row=row, column=2, value=formula)
    c.font = Font(name=FONT, size=10)
    c.number_format = fmt
    c.fill = PatternFill("solid", fgColor=GREY)
    row += 1

row += 1
ws.cell(row=row, column=1, value="NOTE").font = Font(name=FONT, size=10, bold=True, color=NAVY)
row += 1
for line in [
    "Average percentage deviation = mean |scheduled - actual| as a percentage of the "
    f"{CAPACITY_MW} MW installed capacity.",
    "Every metric above is a live formula reading the Day Schedule sheet.",
    "Blocks with no meter reading, or with interpolated (gap-filled) readings, are excluded "
    "from scoring.",
]:
    ws.cell(row=row, column=1, value=line).font = Font(name=FONT, size=9, color="595959")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14

# ============================ METHODOLOGY ============================
ws = wb.create_sheet("Methodology")
title_block(ws, "How this schedule was generated",
            "Per 'Schedule Generation Workflow for Model Evaluation'", 2)

steps = [
    ("Objective",
     "Generate schedules exactly as they would have been generated in real time, then "
     "compare the final schedule against actual meter data."),
    ("1. Windy forecast",
     "Windy satellite video is captured and stored in the S3 bucket at each of the seven "
     "scheduling times. Each scheduling run uses ONLY the clip stored at that time."),
    ("2. Meter data up to block T",
     "Each run receives meter data only up to the current block T. No data from later in "
     "the day is visible to the model - there is no lookahead."),
    ("3. Other inputs",
     "Historical generation, clear-sky physics (pvlib), and the weather forecast (ECMWF, "
     "delivered via Open-Meteo) as it stood at that moment."),
    ("4. Schedule from block T",
     "Using only information available at that instant, the model generates the schedule "
     "from block T to the end of the day."),
    ("5. Next scheduling interval",
     "The next run uses the Windy clip stored for that time and updated meter data up to "
     "the new block T, and regenerates the schedule from that block onward."),
    ("6. Past blocks frozen",
     "Blocks that have already passed keep the values assigned by earlier runs. Only the "
     "remaining future blocks are updated."),
    ("7. Final schedule",
     "Repeating this for every scheduling time yields one complete reconstructed schedule "
     "for the day, closely representing real-time system behaviour."),
    ("8. Evaluation",
     "The final schedule is compared against actual meter data using average percentage "
     "deviation, MAE and RMSE. Only genuinely measured blocks are scored."),
]

row = 4
ws.cell(row=row, column=1, value="STEP").font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
ws.cell(row=row, column=2, value="DESCRIPTION").font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
style_header(ws, row, 2)
row += 1

for i, (step, text) in enumerate(steps):
    a = ws.cell(row=row, column=1, value=step)
    b = ws.cell(row=row, column=2, value=text)
    a.font = Font(name=FONT, size=10, bold=True)
    b.font = Font(name=FONT, size=10)
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b.alignment = Alignment(vertical="top", wrap_text=True)
    a.border = b.border = box
    if i % 2 == 0:
        a.fill = b.fill = PatternFill("solid", fgColor=GREY)
    ws.row_dimensions[row].height = 42
    row += 1

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 95

# =========================== DAY SCHEDULE ============================
ws = wb.create_sheet("Day Schedule")
title_block(ws, f"Reconstructed day schedule - {DAY}",
            "All power values in MW | 15-minute blocks (Indian 96-block convention)", 8)

cols = [
    ("block", "Block No.", 10, "0"),
    ("block_time", "Block Time", 12, None),
    ("scheduled_mw", "Scheduled (MW)", 15, "0.0000"),
    ("actual_mw", "Actual (MW)", 14, "0.0000"),
    ("error_mw", "Error (MW)", 13, "0.0000;-0.0000"),
    ("capacity_utilisation_pct", "Capacity Used (%)", 16, "0.00"),
    ("actual_is_real", "Real Reading", 13, None),
    ("scheduled_at", "Scheduled At", 14, None),
    ("windy_video", "Windy Clip Used", 36, None),
]

header_row = 3
for i, (_, label, width, _fmt) in enumerate(cols, start=1):
    ws.cell(row=header_row, column=i, value=label)
    ws.column_dimensions[get_column_letter(i)].width = width
style_header(ws, header_row, len(cols))

for r, (_, rec) in enumerate(schedule.iterrows(), start=header_row + 1):
    for i, (key, _label, _w, fmt) in enumerate(cols, start=1):
        value = rec[key]
        if pd.isna(value):
            value = None
        if key == "actual_is_real":
            value = bool(value) if value is not None else False
        cell = ws.cell(row=r, column=i, value=value)
        cell.font = Font(name=FONT, size=10)
        cell.border = box
        cell.alignment = Alignment(horizontal="center")
        if fmt:
            cell.number_format = fmt
    if r % 2 == 0:
        for i in range(1, len(cols) + 1):
            ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=GREY)

ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

# ========================= SCHEDULING RUNS ===========================
ws = wb.create_sheet("Scheduling Runs")
title_block(ws, "How the day was built",
            "Each scheduling time rewrites only the blocks still in the future", 5)

headers = ["Scheduling Time", "Windy Clip Used", "Vision Signal",
           "Blocks Written", "Weather Bias Factor"]
widths = [18, 38, 15, 16, 20]

for i, (h, w) in enumerate(zip(headers, widths), start=1):
    ws.cell(row=3, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header(ws, 3, len(headers))

for r, (_, rec) in enumerate(run_log.iterrows(), start=4):
    values = [rec["scheduling_time"], rec["windy_video_used"], rec["vision_signal"],
              int(rec["blocks_written"]), float(rec["weather_bias_factor"])]
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = Font(name=FONT, size=10)
        cell.border = box
        cell.alignment = Alignment(horizontal="center")
        if i == 5:
            cell.number_format = "0.000"
    if r % 2 == 0:
        for i in range(1, len(headers) + 1):
            ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=GREY)

note_row = 4 + len(run_log) + 1
ws.cell(row=note_row, column=1,
        value="Weather Bias Factor: the weather forecast is scaled by this factor, learned "
              "from how much it over-forecast sunlight over the preceding days "
              "(0.732 = the forecast was discounted by 26.8%).")
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
ws.cell(row=note_row, column=1).font = Font(name=FONT, size=9, italic=True, color="595959")
ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[note_row].height = 30

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"written: {OUT}")
